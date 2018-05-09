#! python
# -*- coding:utf-8 -*-

import os
import json
import error
import openpyxl
from slpp.slpp import slpp as lua

# 数组模式下，各个栏的分布
ACMT_ROW = 1  # comment row 注释
ATPE_ROW = 2  # type    row 类型
ASRV_ROW = 3  # server  row 服务器
ACLT_ROW = 4  # client  row 客户端
AKEY_COL = 1  # key  column key所在列

# kv模式下，各个栏的分布
OCMT_COL = 1  # comment column 注释
OTPE_COL = 2  # type    column 类型
OSRV_COL = 3  # server  column 服务器
OCLT_COL = 4  # client  column 客户端
OCTX_COL = 5  # content column 内容所在列
OFLG_ROW = 1  # flag    row    server client所在行

SRV_FLAG = "server"
CLT_FLAG = "client"

SHEET_FLAG_ROW = 1
SHEET_FLAG_COL = 1

ARRAY_FLAG  = "array"
OBJECT_FLAG = "object"

TYPES = { "int":1,"number":2,"int64":3,"string":4,"json":5 }

try:
    basestring
except NameError:
    basestring = str

try:
    long
except NameError:
    long = int

# 类型转换器
class ValueConverter(object):
    def __init__(self):
        pass

    # 在python中，字符串和unicode是不一样的。默认从excel读取的数据都是unicode。
    # str可以通过decode转换为unicode
    # ascii' codec can't encode characters
    def to_unicode_str(self,val):
        if isinstance( val,str ) :
            return val
        else :
            return str( val ).decode("utf8")

    def to_value(self,val_type,val):
        if "int" == val_type :
            return int( val )
        elif "int64" == val_type :
            # 两次转换保证为数字
            return long( val )
        elif "number" == val_type :
            # 去除带小数时的小数点，100.0 ==>> 100
            if long( val ) == float( val ) : return long( val )
            return float( val )
        elif "string" == val_type :
            return self.to_unicode_str( val )
        elif "json" == val_type :
            return json.loads( val )
        else :
            raise Exception( "invalid type",val_type )

# 继承object类，以解决在python2中的错误：TypeError: must be type, not classobj
class Sheet(object):

    def __init__(self,base_name,wb_sheet,srv_writer,clt_writer):
        self.rows   = []
        self.types  = []# 记录各列字段的类型

        self.srv_writer = srv_writer
        self.clt_writer = clt_writer

        self.srv_fields = []#服务端各列字段名
        self.clt_fields = []#客户端各列字段名

        self.converter = ValueConverter()

        self.wb_sheet  = wb_sheet
        self.base_name = base_name

    def write_one_file(self,fields,base_path,writer):
        if len( fields ) <= 0 : return

        wt = writer.Writer( self.base_name,
            self.wb_sheet.title,self.row_offset,self.col_offset )
        ctx = self.writer_content( wt,fields )
        suffix = wt.suffix()

        #必须为wb，不然无法写入utf-8
        path = base_path + self.base_name + "_" + self.wb_sheet.title + suffix
        file = open( path, 'wb' )
        file.write( ctx.encode( "utf-8" ) )
        file.close()

    def write_files(self,srv_path,clt_path):
        pass
        # if None != srv_path and None != self.srv_writer :
        #     self.write_one_file( self.srv_fields,srv_path,self.srv_writer )
        # if None != clt_path and None != self.clt_writer :
        #     self.write_one_file( self.clt_fields,clt_path,self.clt_writer )

# 导出数组类型配置，A1格子的内容有array标识
class ArraySheet(Sheet):

    def __init__(self,base_name,wb_sheet,srv_writer,clt_writer):
        # 记录导出各行的内容
        self.srv_ctx = []
        self.clt_ctx = []

        # 导出的内容起始行数、列数，当发错误时定位
        self.row_offset = ACLT_ROW
        self.col_offset = AKEY_COL
        super( ArraySheet, self ).__init__(
            base_name,wb_sheet,srv_writer,clt_writer )

    # 解析各列的类型(string、number...)
    def decode_type(self):
        # 第一列没数据，类型可以不填，默认为None，但是这里要占个位
        self.types.append( None )

        for col_index in range( AKEY_COL + 1,self.wb_sheet.max_column + 1 ):
            value = self.wb_sheet.cell( row = ATPE_ROW, column = col_index ).value

            # 单元格为空的时候，wb_sheet.cell(row=1, column=2).value == None
            # 类型那一行必须连续，空白表示后面的数据都不导出了
            if value == None: break
            if not TYPES[value]:
                raise Exception( "invalid type",value )

            self.types.append( value )

    # 解析客户端、服务器的字段名(server、client)那两行
    def decode_field(self,fields,row_index):
        for col_index in range( AKEY_COL,len( self.types ) + 1 ):
            value = self.wb_sheet.cell(
                row = row_index, column = col_index ).value

            # 对于不需要导出的field，可以为空。即value为None
            fields.append( value )

    # 解析出一个格子的内容
    def decode_cell(self,row_idx,col_idx):
        value = self.wb_sheet.cell( row = row_idx, column = col_idx ).value
        if not value: return None

        # 类型是从0下标开始，但是excel的第一列从1开始
        return self.converter.to_value( self.types[col_idx - 1],value )

    # 解析出一行的内容
    def decode_row(self,row_idx):
        srv_row = {}
        clt_row = {}

        # 第一列没数据，从第二列开始解析
        for col_idx in range( AKEY_COL + 1,len( self.types ) ):
            value = self.decode_cell( row_idx,col_idx )
            if not value : continue

            srv_key = self.srv_fields[col_idx]
            clt_key = self.clt_fields[col_idx]

            if srv_key : srv_row[srv_key] = value
            if clt_key : clt_row[clt_key] = value

        return srv_row,clt_row # 返回一个tuple

    # 解析导出的内容
    def decode_ctx(self):
        for row_idx in range( ACLT_ROW + 1,self.wb_sheet.max_row + 1 ):
            srv_row,clt_row = self.decode_row( row_idx )

            # 不为空才追加
            if any( srv_row ) : self.srv_ctx.append( srv_row )
            if any( clt_row ) : self.clt_ctx.append( clt_row )

    def writer_content(self,writer,fields):
        return writer.array_content( self.types,fields,self.rows )

    def decode_sheet(self):
        wb_sheet = self.wb_sheet

        self.decode_type()
        if len( self.types ) <= ATPE_ROW:
            print( "    decode sheet %s nothing to decode,abort" \
            % wb_sheet.title.ljust(24,".") )
            return False

        self.decode_field( self.srv_fields,ASRV_ROW )
        self.decode_field( self.clt_fields,ACLT_ROW )

        self.decode_ctx()
        print( self.srv_ctx )
        print( self.clt_ctx )

        print( "    decode sheet %s done" % wb_sheet.title.ljust(24,".") )
        return True

class ObjectSheet(Sheet):

    def __init__(self,base_name,wb_sheet,srv_writer,clt_writer):
        self.row_offset = OCLT_COL
        self.col_offset = OFLG_ROW
        super( ObjectSheet, self ).__init__(
            base_name,wb_sheet,srv_writer,clt_writer )

    def decode_type(self):
        for row_index in range( OFLG_ROW + 1,self.wb_sheet.max_row + 1 ):
            value = self.wb_sheet.cell(
                row = row_index, column = OTPE_COL ).value

            # 单元格为空的时候，wb_sheet.cell(row=1, column=2).value == None
            if value == None: break
            if not TYPES[value]:
                raise Exception( "invalid type",value )

            self.types.append( value )

    def decode_field(self,fields,col_index):
        for row_index in range( OFLG_ROW + 1,len( self.types ) + 2 ):
            value = self.wb_sheet.cell(
                row = row_index, column = col_index ).value

            # 对于不需要导出的field，可以为空。即value为None
            fields.append( value )

    def decode_cell(self):
        # 第一行为flag行，包括最后一行，所以要types + 2
        for row_index in range( OFLG_ROW + 1,len( self.types ) + 2 ):
            value = self.wb_sheet.cell(
                row = row_index, column = OCTX_COL ).value

            self.rows.append( value )

    def writer_content(self,writer,fields):
        return writer.object_content( self.types,fields,self.rows )

    def decode_sheet(self):
        wb_sheet = self.wb_sheet

        self.decode_type()
        if len( self.types ) <= 0:
            print( "    decode sheet %s nothing to decode,abort" \
            % wb_sheet.title.ljust(24,".") )
            return False

        self.decode_field( self.srv_fields,OSRV_COL )
        self.decode_field( self.clt_fields,OCLT_COL )

        self.decode_cell()

        print( "    decode sheet %s done" % wb_sheet.title.ljust(24,".") )
        return True

class ExcelDoc:

    def __init__(self, file,abspath):
        self.file = file
        self.abspath = abspath

    # 是否需要解析
    # 返回解析的对象类型
    def need_decode(self,wb_sheet):
        sheet_val = wb_sheet.cell(
            row = SHEET_FLAG_ROW, column = SHEET_FLAG_ROW ).value

        sheeter = None
        srv_value = None
        clt_value = None
        if ARRAY_FLAG == sheet_val :
            if wb_sheet.max_row <= ACLT_ROW or wb_sheet.max_column <= AKEY_COL:
                return None

            sheeter = ArraySheet
            srv_value = wb_sheet.cell( row = ASRV_ROW, column = AKEY_COL ).value
            clt_value = wb_sheet.cell( row = ACLT_ROW, column = AKEY_COL ).value
        elif OBJECT_FLAG == sheet_val :
            sheeter = ObjectSheet
            srv_value = wb_sheet.cell( row = OFLG_ROW, column = OSRV_COL ).value
            clt_value = wb_sheet.cell( row = OFLG_ROW, column = OCLT_COL ).value
        else :
            return None

        # 没有这两个标识，说明不是配置表。可能是策划的一些备注说明
        if SRV_FLAG != srv_value or CLT_FLAG != clt_value: return None
        return sheeter

    def decode(self,srv_path,clt_path,srv_writer,clt_writer):
        print( "start decode %s ..." % self.file )

        base_name = os.path.splitext( self.file )[0]  #去除后缀
        wb = openpyxl.load_workbook( self.abspath )

        for wb_sheet in wb.worksheets:
            Sheeter = self.need_decode( wb_sheet )

            if None == Sheeter :
                print( "    decode sheet %s no need to decode,abort" % wb_sheet.title.ljust(24,".") )
                continue

            sheet = Sheeter( base_name,wb_sheet,srv_writer,clt_writer )
            if sheet.decode_sheet(): sheet.write_files( srv_path,clt_path )